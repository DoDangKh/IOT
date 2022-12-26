import pickle
import traceback
from datetime import datetime
from email.message import EmailMessage
from tkinter import messagebox
import ssl
import pandas as pd
import openpyxl
from sklearn.tree import DecisionTreeClassifier
email_sender='n19dccn093@student.ptithcm.edu.vn'
email_password='kcxmbelvpalffeaf'
email_receiver='ddangkhoa75@gmail.com'

def create_report_file():
    now=datetime.now()
    day=now.strftime("%A")

    a=[[]]
    for i in range(0,24):
        # print(i)
        # print(day)
        a.append([day,i])
        #print(a)
    del a[0]
    df=pd.DataFrame(a,columns=['Day','Hour'])
    df['Day'].replace(['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday'],
                      [2, 3, 4, 5, 6, 7, 1], inplace=True)
    print(df['Day'][0])

    load_model=pickle.load(open('AI.h5','rb'))
    result=load_model.predict(df)
    df['Day'].replace([1, 2, 3, 4, 5, 6, 7],
                      ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday'],
                      inplace=True)
    wb =  openpyxl.Workbook()
    ws=wb['Sheet']
    for i in range(0,24):
        ws.cell(row=i+1,column=1).value=df['Day'][i]
        ws.cell(row=i+1,column=2).value=df['Hour'][i]
        if(result[i]==1):
            ws.cell(row=i+1,column=3).value="ket xe"
        else:
            ws.cell(row=i + 1, column=3).value = "khong ket xe"
    wb.save("predict.xlsx")

def send_mail():
    import smtplib
    now = datetime.now()
    day = now.strftime("%A")

    a = [[]]
    for i in range(0, 24):
        # print(i)
        # print(day)
        a.append([day, i])
        # print(a)
    del a[0]
    df = pd.DataFrame(a, columns=['Day', 'Hour'])
    df['Day'].replace(['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday'],
                      [2, 3, 4, 5, 6, 7, 1], inplace=True)
    print(df['Day'][0])

    load_model = pickle.load(open('AI.h5', 'rb'))
    result = load_model.predict(df)
    df['Day'].replace([1, 2, 3, 4, 5, 6, 7],
                      ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday'],
                      inplace=True)
    res=''
    print(res)
    for i in range(1,24):
        if(result[i]==1):
            res+= str(i)
            res+="h, "
    print(res)
    try:
        subject='Du Doan Ket Xe'
        em=EmailMessage()
        em['From']=email_sender
        em['To']=email_receiver
        em['Subject']=subject
        em.set_content("Du doan ket xe luc "+res)
        context=ssl.create_default_context()
        with smtplib.SMTP_SSL('smtp.gmail.com',465,context=context) as smtp:
            smtp.login(email_sender,email_password)
            smtp.sendmail(email_sender,email_receiver,em.as_string())
        messagebox.showinfo("Mail","Gui mail thanh cong")
    except:
        traceback.print_exc()
        messagebox.showinfo("Mail","Gui mail that bai")

def train():
    try:
        df=pd.read_excel("data.xlsx")
        df=df.drop(columns=['Sound','Dust'])
        df['Day'].replace(['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday'],
                          [1, 2, 3, 4, 5, 6, 7], inplace=True)
        df['Traffic'].replace(['empty', 'stuck'], [0, 1], inplace=True)
        y = df['Traffic'].values.reshape(-1, 1)
        x = df.drop(columns=['Traffic'])
        my_tree = DecisionTreeClassifier()
        my_tree.fit(x, y)
        pickle.dump(my_tree, open('AI.h5', 'wb'))
        print(df)
        messagebox.showinfo("AI","Train Thanh Cong")
    except:
        messagebox.showinfo("AI", "Train That Bai")

