import os.path
from calendar import weekday
from datetime import datetime

import openpyxl
from openpyxl import load_workbook


def writedata(sound,dust):
    now=datetime.now()
    current=now.strftime("%H")
    day=now.strftime("%A")
    # ws=openpyxl.Workbook.create_sheet
    if(os.path.exists('data.xlsx')):
        wb= load_workbook(filename=r"data.xlsx")

    else:
        wb=openpyxl.Workbook()
        ws = wb["Sheet"]
        ws.cell(row=1, column=1).value = 'Day'
        ws.cell(row=1, column=2).value = 'Hour'
        ws.cell(row=1, column=3).value = 'Sound'
        ws.cell(row=1, column=4).value =  'Dust'
        ws.cell(row=1, column=5).value = 'Traffic'
        #wb.create_sheet("Sheet1")
    ws = wb["Sheet"]
    print(wb.sheetnames)
    lastrow=ws.max_row+1

    ws.cell(row=lastrow,column=1).value=day
    ws.cell(row=lastrow,column=2).value=current
    ws.cell(row=lastrow,column=3).value=sound
    ws.cell(row=lastrow,column=4).value=dust
    if(sound>120 and dust>180):
        ws.cell(row=lastrow,column=5).value="stuck"
    else:
        ws.cell(row=lastrow,column=5).value="empty"
    print(lastrow)
    wb.save("data.xlsx")